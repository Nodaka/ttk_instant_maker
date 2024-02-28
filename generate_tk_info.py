import ast
import copy
import openpyxl
import os
import subprocess
import tempfile
import sys

class class_output:

    def __init__(self):
        self.func_list = []
        self.tkvar_list = []
        self.class_name = None

def class_var_ast(filename):
    if os.path.splitext(filename)[1] == '.py':
        with open(filename) as f:
            base_p = copy.deepcopy(ast.parse(f.read()))
    elif os.path.splitext(filename)[1] == '.ipynb':
        with tempfile.TemporaryDirectory() as td:
            jupyter_path = sys.base_prefix + '\\Scripts\\jupyter.exe'
            instant_py_file_path = td + '\\' + 'instant.py'
            nbconvert_cmd = jupyter_path + ' nbconvert ' + filename + ' --to python --output ' + instant_py_file_path
            returncode = subprocess.Popen(nbconvert_cmd, shell=True)
            _ = returncode.wait()
            with open(instant_py_file_path) as f:
                base_p = copy.deepcopy(ast.parse(f.read()))
    else:
        print('Not python file')
    class_output_list = []
    for node in base_p.body:
        if type(node) == ast.ClassDef:
            instant_class_output = class_output()
            instant_class_output.class_name = node.name
            for node_u in node.body:
                if type(node_u) == ast.FunctionDef:
                    if node_u.name == '__init__':
                        for node_u_u in node_u.body:
                            if type(node_u_u) == ast.Assign:
                                try:
                                    if node_u_u.value.func.attr in ['StringVar', 'IntVar', 'DoubleVar', 'BooleanVar']:
                                        instant_class_output.tkvar_list.append(node_u_u.targets[0].attr)
                                except:
                                    pass
                    else:
                        instant_class_output.func_list.append(node_u.name)
            class_output_list.append(instant_class_output)
    excel_path = '_tk_info.xlsx'
    wb = openpyxl.Workbook()
    sheet_idx = 1
    wb.create_sheet(index=sheet_idx, title='class_name')
    sheet = wb.worksheets[sheet_idx]
    for instant_class_output in class_output_list:
        sheet.append([instant_class_output.class_name])
    sheet_idx = sheet_idx + 1
    for (idx, instant_class_output) in enumerate(class_output_list):
        wb.create_sheet(index=2 * (idx + 1), title=instant_class_output.class_name + '_f')
        sheet = wb.worksheets[2 * (idx + 1)]
        for func_name in instant_class_output.func_list:
            sheet.append([func_name])
        wb.create_sheet(index=2 * (idx + 1) + 1, title=instant_class_output.class_name + '_v')
        sheet = wb.worksheets[2 * (idx + 1) + 1]
        for tkvar_name in instant_class_output.tkvar_list:
            sheet.append([tkvar_name])
    wb.save(excel_path)
    wb.close()